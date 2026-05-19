"""
Parse file Excel TOCFL 華語八千詞 → JSON sạch
Output: tocfl_vocabulary.json
"""
import json
import re
import openpyxl
from pathlib import Path

INPUT_FILE = Path('華語八千詞(內含注音字型檔)/華語八千詞表20240923.xlsx')
OUTPUT_FILE = Path('data/tocfl_vocabulary.json')
OUTPUT_BY_LEVEL = Path('data/tocfl_by_level.json')
OUTPUT_CSV = Path('data/tocfl_vocabulary.csv')

# Map level names
LEVEL_MAP = {
    '準備級一級(Novice 1)': {'code': 'N1', 'name': '準備級一級', 'name_en': 'Novice 1', 'order': 1},
    '準備級二級(Novice 2)': {'code': 'N2', 'name': '準備級二級', 'name_en': 'Novice 2', 'order': 2},
    '入門級(Level 1)':      {'code': 'A1', 'name': '入門級', 'name_en': 'Level 1', 'order': 3},
    '基礎級(Level 2)':      {'code': 'A2', 'name': '基礎級', 'name_en': 'Level 2', 'order': 4},
    '進階級(Level 3)':      {'code': 'B1', 'name': '進階級', 'name_en': 'Level 3', 'order': 5},
    '高階級(Level 4)':      {'code': 'B2', 'name': '高階級', 'name_en': 'Level 4', 'order': 6},
    '流利級(Level 5)':      {'code': 'C1', 'name': '流利級', 'name_en': 'Level 5', 'order': 7},
}

# Map TOCFL level abbreviations in comparison sheet
TOCFL_ABBREV = {
    '準1': 'N1', '準2': 'N2', '入': 'A1', '基': 'A2',
    '進': 'B1', '高': 'B2', '流': 'C1',
}

def clean_text(val):
    """Clean cell value: strip whitespace, handle None"""
    if val is None:
        return None
    return str(val).strip()

def parse_pinyin(raw):
    """Clean pinyin - remove trailing spaces, normalize"""
    if not raw:
        return None
    return raw.strip()

def parse_pos(raw):
    """Parse Parts of Speech - split combined POS like 'N / Vst'"""
    if not raw:
        return []
    # Split by / and clean
    parts = [p.strip() for p in str(raw).split('/')]
    return [p for p in parts if p]

def parse_vocab_sheet(ws, sheet_name, has_context=True):
    """Parse a vocabulary sheet, return list of word dicts"""
    level_info = LEVEL_MAP[sheet_name]
    words = []
    
    for row_idx in range(2, ws.max_row + 1):
        if has_context:
            context = clean_text(ws.cell(row=row_idx, column=1).value)
            vocab = clean_text(ws.cell(row=row_idx, column=2).value)
            pinyin = parse_pinyin(ws.cell(row=row_idx, column=3).value)
            pos_raw = clean_text(ws.cell(row=row_idx, column=4).value)
        else:
            context = None
            vocab = clean_text(ws.cell(row=row_idx, column=1).value)
            pinyin = parse_pinyin(ws.cell(row=row_idx, column=2).value)
            pos_raw = clean_text(ws.cell(row=row_idx, column=3).value)
        
        # Skip empty rows
        if not vocab:
            continue
        
        # Handle words with variants (你/妳, 他/她)
        variants = []
        if '/' in vocab:
            parts = vocab.split('/')
            vocab = parts[0]  # main form
            variants = [p.strip() for p in parts[1:]]
        
        word = {
            'id': f"{level_info['code']}_{len(words) + 1:04d}",
            'vocabulary': vocab,
            'variants': variants,
            'pinyin': pinyin,
            'pos': parse_pos(pos_raw),
            'pos_raw': pos_raw,
            'context': context,
            'level_code': level_info['code'],
            'level_name': level_info['name'],
            'level_name_en': level_info['name_en'],
            'level_order': level_info['order'],
        }
        words.append(word)
    
    return words

def parse_comparison_sheet(ws):
    """Parse the cross-strait comparison sheet"""
    comparisons = []
    for row_idx in range(2, ws.max_row + 1):
        tocfl_word = clean_text(ws.cell(row=row_idx, column=1).value)
        tocfl_level_raw = clean_text(ws.cell(row=row_idx, column=2).value)
        cross_strait = clean_text(ws.cell(row=row_idx, column=3).value)
        hsk_word = clean_text(ws.cell(row=row_idx, column=4).value)
        hsk_level = clean_text(ws.cell(row=row_idx, column=5).value)
        
        if not tocfl_word:
            continue
        
        comparisons.append({
            'tocfl_word': tocfl_word,
            'tocfl_level': TOCFL_ABBREV.get(tocfl_level_raw, tocfl_level_raw),
            'cross_strait_variants': cross_strait,
            'hsk_word': hsk_word,
            'hsk_level': hsk_level,
        })
    return comparisons

def main():
    wb = openpyxl.load_workbook(INPUT_FILE)
    
    all_words = []
    by_level = {}
    
    # Parse vocabulary sheets
    vocab_sheets = [
        ('準備級一級(Novice 1)', True),
        ('準備級二級(Novice 2)', True),
        ('入門級(Level 1)', True),
        ('基礎級(Level 2)', True),
        ('進階級(Level 3)', False),
        ('高階級(Level 4)', False),
        ('流利級(Level 5)', False),
    ]
    
    for sheet_name, has_context in vocab_sheets:
        ws = wb[sheet_name]
        words = parse_vocab_sheet(ws, sheet_name, has_context)
        all_words.extend(words)
        
        level_code = LEVEL_MAP[sheet_name]['code']
        by_level[level_code] = {
            'level_name': LEVEL_MAP[sheet_name]['name'],
            'level_name_en': LEVEL_MAP[sheet_name]['name_en'],
            'order': LEVEL_MAP[sheet_name]['order'],
            'word_count': len(words),
            'words': words,
        }
        print(f"  {sheet_name}: {len(words)} words")
    
    # Parse comparison sheet
    ws_comp = wb['兩岸常用詞語差異表']
    comparisons = parse_comparison_sheet(ws_comp)
    print(f"  Cross-strait comparison: {len(comparisons)} entries")
    
    # Save all words flat
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        json.dump(all_words, f, ensure_ascii=False, indent=2)
    print(f"\n✓ Saved {len(all_words)} words to {OUTPUT_FILE}")
    
    # Save by level
    output_by_level = {
        'total_words': len(all_words),
        'levels': by_level,
        'cross_strait_comparisons': comparisons,
    }
    with open(OUTPUT_BY_LEVEL, 'w', encoding='utf-8') as f:
        json.dump(output_by_level, f, ensure_ascii=False, indent=2)
    print(f"✓ Saved by-level data to {OUTPUT_BY_LEVEL}")
    
    # Save CSV
    import csv
    with open(OUTPUT_CSV, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['id', 'vocabulary', 'variants', 'pinyin', 'pos', 'context', 
                         'level_code', 'level_name', 'level_name_en', 'level_order'])
        for w in all_words:
            writer.writerow([
                w['id'], w['vocabulary'], '|'.join(w['variants']), w['pinyin'],
                '|'.join(w['pos']), w['context'],
                w['level_code'], w['level_name'], w['level_name_en'], w['level_order']
            ])
    print(f"✓ Saved CSV to {OUTPUT_CSV}")
    
    # Summary
    print(f"\n{'='*50}")
    print(f"TOTAL: {len(all_words)} words across {len(by_level)} levels")
    for code, data in sorted(by_level.items(), key=lambda x: x[1]['order']):
        print(f"  {code} {data['level_name']}: {data['word_count']} words")

if __name__ == '__main__':
    main()
