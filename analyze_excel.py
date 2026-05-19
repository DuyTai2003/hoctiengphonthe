import openpyxl

wb = openpyxl.load_workbook('華語八千詞(內含注音字型檔)/華語八千詞表20240923.xlsx')
print('Sheet names:', wb.sheetnames)
for name in wb.sheetnames:
    ws = wb[name]
    print(f'Sheet "{name}": rows={ws.max_row}, cols={ws.max_column}')
    print('  First 3 rows:')
    for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
        print('  ', row)
